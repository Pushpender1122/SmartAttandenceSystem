import face_recognition
import cv2
import pickle
import openpyxl
from datetime import datetime
import requests

# Load or create an Excel sheet for attendance tracking
def load_or_create_workbook(filename="attendance.xlsx"):
    try:
        # Try to load the existing workbook
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # Create a new workbook and sheet if it doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        # Add headers to the Excel sheet
        sheet.append(["Student ID", "Date", "Status"])
        workbook.save(filename)
    return workbook

# Function to log attendance
def log_attendance(image_id):
    workbook = load_or_create_workbook()
    sheet = workbook["Attendance"]
    
    # Get the current date
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # Check if the student is already marked present for today
    for row in sheet.iter_rows(values_only=True):
        if row[0] == image_id and row[1] == current_date:
            print(f"Attendance already recorded for ID: {image_id} on {current_date}")
            return  # If attendance is already recorded for today, skip

    # Log the student's attendance (ID, Date, Status)
    sheet.append([image_id, current_date, "Present"])
    workbook.save("attendance.xlsx")
    print(f"Attendance recorded for ID: {image_id}")

    # Send the data to Node.js
    url = "https://smartattandencesystem.onrender.com/add"  # Replace with your Node.js API endpoint
    payload = {"image_id": image_id, "date": current_date, "status": "Present"}
    try:
        response = requests.post(url, json=payload)
        if response.status_code == 200 or response.status_code == 201:
            print("Data sent to Node.js successfully:", response.json())
        else:
            print(f"Failed to send data to Node.js. Status code: {response.status_code}, Response: {response.text}")
    except Exception as e:
        print(f"Error sending data to Node.js: {e}")

def start_attendance(name,whoIs):
        # Load the saved face encodings from the pickle file
    with open(name, "rb") as f:
        saved_face_encodings = pickle.load(f)

    # Initialize the webcam
    video_capture = cv2.VideoCapture(0)
    print("Starting webcam. Press 'q' to quit.")

    while True: 
        # Capture frame from the webcam
        ret, frame = video_capture.read()
        
        if not ret:
            print("Failed to capture frame from webcam. Exiting...")
            break

        # Convert the frame from BGR (OpenCV format) to RGB (required by face_recognition)
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        
        # Find all face locations in the frame
        face_locations = face_recognition.face_locations(rgb_frame)
        
        # Compute the face encodings for the found faces in the frame
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        
        for face_encoding in face_encodings:
            # Compare the current face encoding with the saved ones
            match = False
            for image_id, saved_encoding in saved_face_encodings.items():
                matches = face_recognition.compare_faces([saved_encoding], face_encoding)
                face_distances = face_recognition.face_distance([saved_encoding], face_encoding)
                if matches[0] and face_distances[0] < 0.6:  # You can adjust the threshold as needed
                    match = True
                    if whoIs == "teacher":
                        log_attendance(image_id)
                        video_capture.release()
                        cv2.destroyAllWindows()
                    else:
                        log_attendance(image_id)  # Log attendance when a match is found
                    break
            
            if not match:
                print("No match found.")
        
        # Display the frame with face detection rectangles
        for (top, right, bottom, left) in face_locations:
            # Draw a rectangle around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        
        # Show the frame in a window
        cv2.imshow('Video', frame)

        # Press 'q' to quit the video window
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the webcam and close windows
    video_capture.release()
    cv2.destroyAllWindows()
